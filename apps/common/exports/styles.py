from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CONTENT_TYPE_XLSX = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# ---------------------------------------------------------------------------
# Estilos de encabezado COMPARTIDOS por todos los Excel del sistema:
#   - Export de datos reales  -> ExcelExportMixin
#   - Plantillas de ejemplo   -> BaseExampleExportView
# Cambiar algo aquí se refleja en AMBOS, manteniéndolos alineados.
# Colores de marca (ver static/css/custom/colors.main.css):
#   --color-primary (sand-500) = #D7B28A · texto blanco (#FFFFFF)
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="D7B28A")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
HEADER_HEIGHT = 22
FREEZE_HEADER = True


def style_header_cell(cell):
    """Aplica los estilos de encabezado de marca a una celda."""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGNMENT
    return cell


def write_header_row(sheet, headers, widths=None):
    """Escribe y estiliza la fila 1 de encabezados (estilos de marca).

    headers: lista de textos de encabezado.
    widths:  lista opcional de anchos por columna (mismo orden que headers).
    """
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        style_header_cell(cell)
        if widths:
            width = widths[col_index - 1] if col_index <= len(widths) else None
            if width:
                sheet.column_dimensions[get_column_letter(col_index)].width = width
    sheet.row_dimensions[1].height = HEADER_HEIGHT
    return sheet
